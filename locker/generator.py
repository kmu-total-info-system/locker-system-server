from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.worksheet.pagebreak import Break

sheets_name = ['1학년', '2학년', '3학년', '4학년']

# datas = [
#     {'name': '홍승의', 'user_id': '20171722', 'organization': '제 3대 소프트웨어 융합 학생회 리턴', 'locker': 'A - 1', 'grade': 1},
#     {'name': '홍승의', 'user_id': '20171722', 'organization': '제 3대 소프트웨어 융합 학생회 리턴', 'locker': 'A - 1', 'grade': 1},
#     {'name': '홍승의', 'user_id': '20171722', 'organization': '제 3대 소프트웨어 융합 학생회 리턴', 'locker': 'A - 1', 'grade': 1},
#     {'name': '홍승의', 'user_id': '20171722', 'organization': '제 3대 소프트웨어 융합 학생회 리턴', 'locker': 'A - 1', 'grade': 1},
#     {'name': '홍승의', 'user_id': '20171722', 'organization': '제 3대 소프트웨어 융합 학생회 리턴', 'locker': 'A - 1', 'grade': 1},
#     {'name': '홍승의', 'user_id': '20171722', 'organization': '제 3대 소프트웨어 융합 학생회 리턴', 'locker': 'A - 1', 'grade': 1},
#     {'name': '홍승의', 'user_id': '20171722', 'organization': '제 3대 소프트웨어 융합 학생회 리턴', 'locker': 'A - 1', 'grade': 1},
#     {'name': '홍승의', 'user_id': '20171722', 'organization': '제 3대 소프트웨어 융합 학생회 리턴', 'locker': 'A - 1', 'grade': 1},
#     {'name': '홍승의', 'user_id': '20171722', 'organization': '제 3대 소프트웨어 융합 학생회 리턴', 'locker': 'A - 1', 'grade': 1},
#     {'name': '홍승의', 'user_id': '20171722', 'organization': '제 3대 소프트웨어 융합 학생회 리턴', 'locker': 'A - 1', 'grade': 1},
#     {'name': '홍승의', 'user_id': '20171722', 'organization': '제 3대 소프트웨어 융합 학생회 리턴', 'locker': 'A - 1', 'grade': 1},
#     {'name': '홍승의', 'user_id': '20171722', 'organization': '제 3대 소프트웨어 융합 학생회 리턴', 'locker': 'A - 1', 'grade': 1},
#     {'name': '홍승의', 'user_id': '20171722', 'organization': '제 3대 소프트웨어 융합 학생회 리턴', 'locker': 'A - 1', 'grade': 1},
#     {'name': '홍승의', 'user_id': '20171722', 'organization': '제 3대 소프트웨어 융합 학생회 리턴', 'locker': 'A - 1', 'grade': 1},
#     {'name': '홍승의', 'user_id': '20171722', 'organization': '제 3대 소프트웨어 융합 학생회 리턴', 'locker': 'A - 1', 'grade': 1},
#     {'name': '홍승의', 'user_id': '20171722', 'organization': '제 3대 소프트웨어 융합 학생회 리턴', 'locker': 'A - 1', 'grade': 1},
#     {'name': '홍승의', 'user_id': '20171722', 'organization': '제 3대 소프트웨어 융합 학생회 리턴', 'locker': 'A - 1', 'grade': 1},
#     {'name': '홍승의', 'user_id': '20171722', 'organization': '제 3대 소프트웨어 융합 학생회 리턴', 'locker': 'A - 1', 'grade': 1},
#     {'name': '홍승의', 'user_id': '20171722', 'organization': '제 3대 소프트웨어 융합 학생회 리턴', 'locker': 'A - 1', 'grade': 1},
#     {'name': '홍승의', 'user_id': '20171722', 'organization': '제 3대 소프트웨어 융합 학생회 리턴', 'locker': 'A - 1', 'grade': 1},
#     {'name': '홍승의', 'user_id': '20171722', 'organization': '제 3대 소프트웨어 융합 학생회 리턴', 'locker': 'A - 1', 'grade': 1},
#     {'name': '홍승의', 'user_id': '20171722', 'organization': '제 3대 소프트웨어 융합 학생회 리턴', 'locker': 'A - 1', 'grade': 1},
#     {'name': '홍승의', 'user_id': '20171722', 'organization': '제 3대 소프트웨어 융합 학생회 리턴', 'locker': 'A - 1', 'grade': 1},
#     {'name': '홍승의', 'user_id': '20171722', 'organization': '제 3대 소프트웨어 융합 학생회 리턴', 'locker': 'A - 1', 'grade': 1},
#     {'name': '홍승의', 'user_id': '20171722', 'organization': '제 3대 소프트웨어 융합 학생회 리턴', 'locker': 'A - 1', 'grade': 1},
#     {'name': '홍승의', 'user_id': '20171722', 'organization': '제 3대 소프트웨어 융합 학생회 리턴', 'locker': 'A - 1', 'grade': 1},
#     {'name': '홍승의', 'user_id': '20171722', 'organization': '제 3대 소프트웨어 융합 학생회 리턴', 'locker': 'A - 1', 'grade': 1},
#     {'name': '홍승의', 'user_id': '20171722', 'organization': '제 3대 소프트웨어 융합 학생회 리턴', 'locker': 'A - 1', 'grade': 1},
#     {'name': '홍승의', 'user_id': '20171722', 'organization': '제 3대 소프트웨어 융합 학생회 리턴', 'locker': 'A - 1', 'grade': 1},
#     {'name': '홍승의', 'user_id': '20171722', 'organization': '제 3대 소프트웨어 융합 학생회 리턴', 'locker': 'A - 1', 'grade': 1},
#     {'name': '홍승의', 'user_id': '20171722', 'organization': '제 3대 소프트웨어 융합 학생회 리턴', 'locker': 'A - 1', 'grade': 1},
#     {'name': '홍승의', 'user_id': '20171722', 'organization': '제 3대 소프트웨어 융합 학생회 리턴', 'locker': 'A - 1', 'grade': 1},
#     {'name': '홍승의', 'user_id': '20171722', 'organization': '제 3대 소프트웨어 융합 학생회 리턴', 'locker': 'A - 1', 'grade': 1},
#     {'name': '홍승의', 'user_id': '20171722', 'organization': '제 3대 소프트웨어 융합 학생회 리턴', 'locker': 'A - 1', 'grade': 1},
#     {'name': '홍승의', 'user_id': '20171722', 'organization': '제 3대 소프트웨어 융합 학생회 리턴', 'locker': 'A - 1', 'grade': 1},
#     {'name': '홍승의', 'user_id': '20171722', 'organization': '제 3대 소프트웨어 융합 학생회 리턴', 'locker': 'A - 1', 'grade': 1},
#     {'name': '홍승의', 'user_id': '20171722', 'organization': '제 3대 소프트웨어 융합 학생회 리턴', 'locker': 'A - 1', 'grade': 1},
#     {'name': '홍승의', 'user_id': '20171722', 'organization': '제 3대 소프트웨어 융합 학생회 리턴', 'locker': 'A - 1', 'grade': 1},
#     {'name': '홍승의', 'user_id': '20171722', 'organization': '제 3대 소프트웨어 융합 학생회 리턴', 'locker': 'A - 1', 'grade': 1},
#     {'name': '홍승의', 'user_id': '20171722', 'organization': '제 3대 소프트웨어 융합 학생회 리턴', 'locker': 'A - 1', 'grade': 1},
#     {'name': '홍승의', 'user_id': '20171722', 'organization': '제 3대 소프트웨어 융합 학생회 리턴', 'locker': 'A - 1', 'grade': 1},
#     {'name': '홍승의', 'user_id': '20171722', 'organization': '제 3대 소프트웨어 융합 학생회 리턴', 'locker': 'A - 1', 'grade': 1},
#     {'name': '홍승의', 'user_id': '20171722', 'organization': '제 3대 소프트웨어 융합 학생회 리턴', 'locker': 'A - 1', 'grade': 1},
#     {'name': '홍승의', 'user_id': '20171722', 'organization': '제 3대 소프트웨어 융합 학생회 리턴', 'locker': 'A - 1', 'grade': 1},
#     {'name': '홍승의', 'user_id': '20171722', 'organization': '제 3대 소프트웨어 융합 학생회 리턴', 'locker': 'A - 1', 'grade': 1},
#     {'name': '홍승의', 'user_id': '20171722', 'organization': '제 3대 소프트웨어 융합 학생회 리턴', 'locker': 'A - 1', 'grade': 1},
#     {'name': '홍승의', 'user_id': '20171722', 'organization': '제 3대 소프트웨어 융합 학생회 리턴', 'locker': 'A - 1', 'grade': 1},
#     {'name': '홍승의', 'user_id': '20171722', 'organization': '제 3대 소프트웨어 융합 학생회 리턴', 'locker': 'A - 1', 'grade': 1},
#     {'name': '홍승의', 'user_id': '20171722', 'organization': '제 3대 소프트웨어 융합 학생회 리턴', 'locker': 'A - 1', 'grade': 1},
#     {'name': '홍승의', 'user_id': '20171722', 'organization': '제 3대 소프트웨어 융합 학생회 리턴', 'locker': 'A - 1', 'grade': 1},
#     {'name': '홍승의', 'user_id': '20171722', 'organization': '제 3대 소프트웨어 융합 학생회 리턴', 'locker': 'A - 1', 'grade': 1},
#     {'name': '홍승의', 'user_id': '20171722', 'organization': '제 3대 소프트웨어 융합 학생회 리턴', 'locker': 'A - 1', 'grade': 1},
#
# ]
def is_page_divide(count):
    if (count % 16 == 0 or count % 16 == 1) and count != 0 and count != 1:
        return True
    return False

def generate(datas):
    wb = Workbook(write_only=False)
    ws = [wb.create_sheet(sheet_name) for sheet_name in sheets_name]
    count = [0 for i in sheets_name]

    for row in datas:
        sheet_num = row['grade'] - 1

        if is_page_divide(count[sheet_num]):
            ws[sheet_num].row_breaks.append(Break(id=int(count[sheet_num] / 2) * 6 + 1))

        organization = ws[sheet_num].cell(column=count[sheet_num] % 2 * 5 + 3,
                                          row=int(count[sheet_num] / 2) * 6 + 2,
                                          value=row['organization'])

        locker = ws[sheet_num].cell(column=count[sheet_num] % 2 * 5 + 3,
                                    row=int(count[sheet_num] / 2) * 6 + 4,
                                    value='[ ' + row['locker'] + ' ]')

        user = ws[sheet_num].cell(column=count[sheet_num] % 2 * 5 + 3,
                                  row=int(count[sheet_num] / 2) * 6 + 5,
                                  value=row['user_id'] + ' ' + row['name'])

        count[sheet_num] += 1

        organization.font = Font(size=24, bold=True)
        organization.alignment = Alignment(horizontal='center')
        locker.font = Font(size=24, bold=True)
        locker.alignment = Alignment(horizontal='center')
        user.font = Font(size=24, bold=True)
        user.alignment = Alignment(horizontal='center')

    for i in range(4):
        for c in "ABCDEFGHIJKLMNOPQ":
            ws[i].column_dimensions[c].width = 14.2857142857
        ws[i].page_setup.orientation = ws[i].ORIENTATION_PORTRAIT
        ws[i].page_setup.paperSize = ws[i].PAPERSIZE_A4
        ws[i].sheet_properties.pageSetUpPr.fitToPage = True
        ws[i].page_setup.fitToHeight = False
        ws[i].page_setup.fitToWidth = True
        ws[i].page_margins.left = 0.7
        ws[i].page_margins.top = 0.75
        ws[i].page_margins.right = 0.7
        ws[i].page_margins.bottom = 0.8
        ws[i].sheet_view.showGridLines = False

    return wb



